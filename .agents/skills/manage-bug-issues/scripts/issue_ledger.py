#!/usr/bin/env python3
"""Fail-fast, append-audited XLSX bug-issue ledger.

The workbook is deliberately the only persistence boundary.  This module exposes
domain commands; it does not expose a generic spreadsheet editor.
"""
from __future__ import annotations

import argparse
import contextlib
import datetime as dt
import fcntl
import hashlib
import json
import os
import re
import subprocess
import sys
import tempfile
import uuid
from pathlib import Path, PurePosixPath
from typing import Any, Iterator, NoReturn, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


class LedgerError(Exception):
    """A user-correctable ledger invariant violation."""


SHEETS: dict[str, list[str]] = {
    "Project": ["project_id", "name", "created_at", "version"],
    "Issues": ["issue_id", "project_id", "title", "description", "severity", "status", "version", "latest_event_id", "reported_at", "reported_by", "root_cause", "authorization_scope", "authorized_by", "authorized_at", "resolution", "closed_at", "updated_at"],
    "Events": ["event_id", "issue_id", "event_type", "from_status", "to_status", "occurred_at", "actor", "payload_json", "previous_hash", "event_hash", "corrects_event_id"],
    "Evidence": ["evidence_id", "issue_id", "evidence_type", "repo_id", "commit_hash", "relative_path", "locator", "description", "registered_at", "registered_by", "active"],
    "EventEvidence": ["event_id", "evidence_id"],
    "IssueRelations": ["relation_id", "source_issue_id", "target_issue_id", "relation_type", "created_at", "created_by"],
    "Repositories": ["repo_id", "role", "name", "canonical_url", "default_branch", "active"],
    "Dictionaries": ["dictionary_type", "key", "value", "active"],
}

STATES = ("REPORTED", "TRIAGED", "CONFIRMED", "INVESTIGATING", "READY_TO_FIX", "FIXING", "READY_TO_VERIFY", "CLOSED", "REOPENED", "DEFERRED", "DUPLICATE", "MERGED", "INVALID", "CANNOT_REPRODUCE", "WONT_FIX")
TERMINAL = {"CLOSED", "DEFERRED", "DUPLICATE", "MERGED", "INVALID", "CANNOT_REPRODUCE", "WONT_FIX"}
LEDGER_SCHEMA_VERSION = 2
RELATION_TYPES = {"DUPLICATE_OF", "MERGED_INTO", "REGRESSION_OF", "RELATED_TO"}
CONSOLIDATION_RELATIONS = {"DUPLICATE": "DUPLICATE_OF", "MERGED": "MERGED_INTO"}
CONSOLIDATION_RELATION_TYPES = set(CONSOLIDATION_RELATIONS.values())
IDEMPOTENCY_FIELDS = (
    "issue_id", "event_id", "evidence_id", "title", "description", "severity",
    "actor", "expected_version", "root_cause", "authorized_by",
    "authorization_scope", "resolution", "verification_result",
    "target_issue_id", "relation_type", "evidence_type", "repo_id", "commit",
    "relative_path", "locator", "progress",
)
SELF_EVENT_TYPES = {"IDENTIFY_ROOT_CAUSE", "RECORD_PROGRESS", "RELATED", "EVIDENCE_REGISTERED", "EVIDENCE_RETRACTED", "EVENT_CORRECTED"}
EVENT_TARGETS: dict[str, set[str]] = {
    "RECORD_HISTORY_SEARCH": {"TRIAGED"},
    "CONFIRM": {"CONFIRMED"},
    "START_INVESTIGATION": {"INVESTIGATING"},
    "AUTHORIZE_FIX": {"READY_TO_FIX"},
    "REVOKE_FIX_AUTHORIZATION": {"INVESTIGATING"},
    "START_FIX": {"FIXING"},
    "COMPLETE_FIX": {"READY_TO_VERIFY"},
    "RECORD_VERIFICATION": {"CLOSED", "REOPENED"},
    "REOPEN": {"REOPENED"},
    "DEFER": {"DEFERRED"},
    "RESOLVE_WITHOUT_FIX": {"DUPLICATE", "MERGED", "INVALID", "CANNOT_REPRODUCE", "WONT_FIX"},
}
EVENT_REQUIRED_PAYLOAD: dict[str, set[str]] = {
    "REPORTED": {"title", "description", "severity"},
    "IDENTIFY_ROOT_CAUSE": {"root_cause"},
    "RECORD_PROGRESS": {"progress"},
    "AUTHORIZE_FIX": {"authorized_by", "authorization_scope"},
    "RECORD_VERIFICATION": {"verification_result", "resolution"},
    "DEFER": {"resolution"},
    "RESOLVE_WITHOUT_FIX": {"resolution"},
    "EVIDENCE_REGISTERED": {"evidence_type", "repo_id", "commit_hash", "description"},
    "EVIDENCE_RETRACTED": {"evidence_id"},
    "EVENT_CORRECTED": {"corrected_event_id"},
}
LEGACY_REPORTED_TARGETS = {"DEFERRED", "DUPLICATE", "INVALID", "CANNOT_REPRODUCE"}
TRANSITIONS: dict[str, set[str]] = {
    "REPORTED": {"TRIAGED"},
    "TRIAGED": {"CONFIRMED", "DEFERRED", "DUPLICATE", "MERGED", "INVALID", "CANNOT_REPRODUCE", "WONT_FIX"},
    "CONFIRMED": {"INVESTIGATING", "DEFERRED", "DUPLICATE", "MERGED", "WONT_FIX"},
    "INVESTIGATING": {"READY_TO_FIX", "DEFERRED", "DUPLICATE", "MERGED", "CANNOT_REPRODUCE", "WONT_FIX", "INVALID"},
    "READY_TO_FIX": {"FIXING", "INVESTIGATING", "DEFERRED", "DUPLICATE", "MERGED", "WONT_FIX"},
    "FIXING": {"READY_TO_VERIFY", "INVESTIGATING", "DEFERRED", "DUPLICATE", "MERGED"},
    "READY_TO_VERIFY": {"CLOSED", "REOPENED", "FIXING", "INVESTIGATING", "DEFERRED", "DUPLICATE", "MERGED"},
    "REOPENED": {"CONFIRMED", "INVESTIGATING", "DEFERRED", "DUPLICATE", "MERGED"},
    "CLOSED": {"REOPENED"}, "DEFERRED": {"REOPENED"}, "DUPLICATE": {"REOPENED"},
    "MERGED": {"REOPENED"}, "INVALID": {"REOPENED"}, "CANNOT_REPRODUCE": {"REOPENED"}, "WONT_FIX": {"REOPENED"},
}
SHA = re.compile(r"^[0-9a-f]{40}$")
SHA256 = re.compile(r"^[0-9a-f]{64}$")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TEXT_COLUMNS = {
    "name": 24,
    "title": 34,
    "description": 50,
    "payload_json": 60,
    "root_cause": 50,
    "authorization_scope": 40,
    "resolution": 40,
    "canonical_url": 50,
    "relative_path": 40,
    "locator": 30,
}
FIELD_WIDTHS = {
    "project_id": 14,
    "issue_id": 18,
    "event_id": 20,
    "evidence_id": 20,
    "relation_id": 20,
    "dictionary_type": 22,
    "key": 22,
    "value": 22,
    "status": 20,
    "from_status": 20,
    "to_status": 20,
    "event_type": 24,
    "evidence_type": 22,
    "commit_hash": 20,
    "previous_hash": 20,
    "event_hash": 20,
    "latest_event_id": 20,
    "created_at": 26,
    "reported_at": 26,
    "updated_at": 26,
    "closed_at": 26,
    "authorized_at": 26,
    "occurred_at": 26,
    "registered_at": 26,
}


def now() -> str:
    return dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat()


def fail(message: str) -> NoReturn:
    raise LedgerError(message)


def required(value: Any, name: str) -> Any:
    if value is None or value == "":
        fail(f"missing required field: {name}")
    return value


def json_object(value: str | None, name: str = "payload_json") -> dict[str, Any]:
    if not value:
        return {}
    try:
        result = json.loads(value)
    except json.JSONDecodeError as exc:
        fail(f"invalid {name}: {exc.msg}")
    if not isinstance(result, dict):
        fail(f"{name} must be a JSON object")
    return result


def canonical(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def event_digest(event: dict[str, Any]) -> str:
    body = {k: value for k, value in event.items() if k != "event_hash"}
    return hashlib.sha256(canonical(body).encode()).hexdigest()


def cell(value: Any) -> Any:
    return "" if value is None else value


@contextlib.contextmanager
def workbook_lock(path: Path, exclusive: bool) -> Iterator[None]:
    lock_path = path.with_name(path.name + ".lock")
    with lock_path.open("a+b") as handle:
        fcntl.flock(handle.fileno(), fcntl.LOCK_EX if exclusive else fcntl.LOCK_SH)
        try:
            yield
        finally:
            fcntl.flock(handle.fileno(), fcntl.LOCK_UN)


class Ledger:
    def __init__(self, path: Path, writable: bool):
        self.path, self.writable = path, writable
        if not path.is_file():
            fail(f"workbook does not exist: {path}")
        try:
            self.wb = load_workbook(path)
        except Exception as exc:
            fail(f"cannot open workbook: {exc}")
        self._validate_shape()

    def _validate_shape(self) -> None:
        if set(self.wb.sheetnames) != set(SHEETS):
            fail("workbook sheets do not match the controlled ledger schema")
        for name, columns in SHEETS.items():
            ws = self.wb[name]
            if list(ws.values)[0] != tuple(columns):
                fail(f"{name} headers do not match the controlled ledger schema")
            if name not in ws.tables:
                fail(f"missing controlled table: {name}")

    def rows(self, sheet: str) -> list[dict[str, Any]]:
        ws, headers = self.wb[sheet], SHEETS[sheet]
        result: list[dict[str, Any]] = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in values):
                continue
            if len(values) != len(headers):
                fail(f"malformed row in {sheet}")
            result.append(dict(zip(headers, ("" if v is None else v for v in values))))
        return result

    def append(self, sheet: str, row: dict[str, Any]) -> None:
        headers = SHEETS[sheet]
        if set(row) != set(headers):
            fail(f"internal schema violation while appending {sheet}")
        ws = self.wb[sheet]
        ws.append([cell(row[h]) for h in headers])
        ws.row_dimensions[ws.max_row].height = 30 if any(row.get(h) for h in TEXT_COLUMNS) else 20
        for column, header in enumerate(headers, start=1):
            target = ws.cell(ws.max_row, column)
            target.alignment = Alignment(
                vertical="top",
                wrap_text=header in TEXT_COLUMNS,
            )
            if header.endswith("_at"):
                target.number_format = "yyyy-mm-dd hh:mm"

    def replace(self, sheet: str, key: str, value: str, changes: dict[str, Any]) -> dict[str, Any]:
        ws, headers = self.wb[sheet], SHEETS[sheet]
        if not set(changes).issubset(headers):
            fail("attempted undefined ledger field")
        index = {h: n + 1 for n, h in enumerate(headers)}
        for row_no in range(2, ws.max_row + 1):
            if ws.cell(row_no, index[key]).value == value:
                for field, field_value in changes.items():
                    ws.cell(row_no, index[field]).value = cell(field_value)
                record = {h: cell(ws.cell(row_no, index[h]).value) for h in headers}
                return record
        fail(f"{sheet} record not found: {value}")

    def find(self, sheet: str, key: str, value: str) -> dict[str, Any]:
        found = [r for r in self.rows(sheet) if r[key] == value]
        if len(found) != 1:
            fail(f"expected one {sheet} record for {key}={value}, found {len(found)}")
        return found[0]

    def save(self) -> None:
        for name in SHEETS:
            ws = self.wb[name]
            ws.tables[name].ref = f"A1:{chr(64 + len(SHEETS[name]))}{max(1, ws.max_row)}"
            ws.print_area = ws.tables[name].ref
        fd, temporary = tempfile.mkstemp(prefix=f".{self.path.stem}.", suffix=".xlsx", dir=self.path.parent)
        os.close(fd)
        try:
            self.wb.save(temporary)
            Ledger(Path(temporary), True).validate()
            with open(temporary, "rb") as handle:
                os.fsync(handle.fileno())
            os.replace(temporary, self.path)
        finally:
            if os.path.exists(temporary):
                os.unlink(temporary)

    def validate(self) -> dict[str, Any]:
        self._validate_shape()
        project = self.rows("Project")
        if len(project) != 1 or not project[0]["project_id"]:
            fail("Project must contain exactly one project")
        schema_version = project[0]["version"]
        if schema_version not in {1, LEDGER_SCHEMA_VERSION}: fail(f"unsupported ledger schema version: {schema_version}")
        issues = self.rows("Issues")
        issue_ids = {r["issue_id"] for r in issues}
        if len(issue_ids) != len(issues): fail("duplicate issue_id")
        for issue in issues:
            if issue["project_id"] != project[0]["project_id"]: fail("issue project foreign key violation")
            if issue["status"] not in STATES: fail(f"unknown issue status: {issue['status']}")
            if not isinstance(issue["version"], int) or issue["version"] < 1: fail("invalid issue version")
        repos = {r["repo_id"] for r in self.rows("Repositories")}
        events = self.rows("Events")
        event_ids: set[str] = set(); event_hashes: dict[str, str] = {}; event_payloads: dict[str, dict[str, Any]] = {}; strict_events: dict[str, bool] = {}; idempotency_keys: set[str] = set()
        last_hash: dict[str, str] = {}; last_event: dict[str, str] = {}; last_status: dict[str, str] = {}
        for event in events:
            if event["event_id"] in event_ids: fail("duplicate event_id")
            event_ids.add(event["event_id"])
            if event["issue_id"] not in issue_ids: fail("event issue foreign key violation")
            if event["previous_hash"] != last_hash.get(event["issue_id"], ""): fail("broken event hash chain")
            if event["event_hash"] != event_digest(event): fail("invalid event hash")
            if event["corrects_event_id"] and event["corrects_event_id"] not in event_hashes: fail("correction must reference an earlier event")
            if event["from_status"] not in STATES or event["to_status"] not in STATES: fail("event has invalid status snapshot")
            stored_payload = json_object(event["payload_json"], "stored event payload")
            request_fingerprint = stored_payload.get("request_fingerprint")
            if schema_version >= LEDGER_SCHEMA_VERSION and not request_fingerprint: fail("schema v2 event requires request_fingerprint")
            if request_fingerprint and not SHA256.fullmatch(str(request_fingerprint)): fail("invalid request_fingerprint")
            strict_event = schema_version >= LEDGER_SCHEMA_VERSION or bool(request_fingerprint)
            if event["issue_id"] not in last_status:
                if (event["event_type"], event["from_status"], event["to_status"]) != ("REPORTED", "REPORTED", "REPORTED"): fail("first issue event must be REPORTED -> REPORTED")
            else:
                if event["from_status"] != last_status[event["issue_id"]]: fail("event status snapshots are not continuous")
                if event["event_type"] in SELF_EVENT_TYPES:
                    if event["from_status"] != event["to_status"]: fail(f"{event['event_type']} must not change issue status")
                elif event["event_type"] in EVENT_TARGETS:
                    if event["from_status"] == event["to_status"] or event["to_status"] not in EVENT_TARGETS[event["event_type"]]: fail(f"event type has invalid target status: {event['event_type']} -> {event['to_status']}")
                else:
                    fail(f"unknown or misplaced event_type: {event['event_type']}")
            if strict_event:
                missing_payload = [
                    name for name in EVENT_REQUIRED_PAYLOAD.get(event["event_type"], set())
                    if stored_payload.get(name) in (None, "")
                ]
                if missing_payload: fail(f"event payload is missing required fields: {', '.join(sorted(missing_payload))}")
                if event["event_type"] == "RESOLVE_WITHOUT_FIX" and stored_payload["resolution"] != event["to_status"]:
                    fail("resolution payload does not match terminal status")
                if event["event_type"] == "RECORD_VERIFICATION":
                    verification_target = {"PASS": "CLOSED", "FAIL": "REOPENED"}.get(stored_payload["verification_result"])
                    if verification_target != event["to_status"]: fail("verification_result does not match target status")
            if event["from_status"] != event["to_status"] and event["to_status"] not in TRANSITIONS[event["from_status"]]:
                legacy_reported_transition = not strict_event and event["from_status"] == "REPORTED" and event["to_status"] in LEGACY_REPORTED_TARGETS
                if not legacy_reported_transition: fail(f"event has invalid state transition: {event['from_status']} -> {event['to_status']}")
            event_hashes[event["event_id"]] = event["event_hash"]; last_hash[event["issue_id"]] = event["event_hash"]; last_event[event["issue_id"]] = event["event_id"]; last_status[event["issue_id"]] = event["to_status"]
            idempotency_key = stored_payload.get("idempotency_key")
            if idempotency_key:
                if not isinstance(idempotency_key, str) or idempotency_key in idempotency_keys: fail("duplicate or invalid idempotency_key")
                idempotency_keys.add(idempotency_key)
            event_payloads[event["event_id"]] = stored_payload
            strict_events[event["event_id"]] = strict_event
        evidence_ids: set[str] = set()
        for evidence in self.rows("Evidence"):
            if evidence["evidence_id"] in evidence_ids: fail("duplicate evidence_id")
            evidence_ids.add(evidence["evidence_id"])
            if evidence["issue_id"] not in issue_ids or evidence["repo_id"] not in repos or not evidence["evidence_type"]: fail("evidence foreign key violation")
            if not SHA.fullmatch(str(evidence["commit_hash"])): fail("evidence commit is not a canonical full SHA-1")
            self._relative(evidence["relative_path"])
            if evidence["active"] not in (True, False, 0, 1): fail("invalid evidence active value")
        for issue in issues:
            if issue["latest_event_id"] != last_event.get(issue["issue_id"], ""): fail("issue snapshot latest_event_id is inconsistent")
            if issue["status"] != last_status.get(issue["issue_id"], ""): fail("issue snapshot status is inconsistent with latest event")
        for link in self.rows("EventEvidence"):
            if link["event_id"] not in event_ids or link["evidence_id"] not in evidence_ids: fail("EventEvidence foreign key violation")
        relations = self.rows("IssueRelations")
        relation_ids: set[str] = set(); relations_by_id: dict[str, dict[str, Any]] = {}
        for relation in relations:
            if relation["relation_id"] in relation_ids: fail("duplicate relation_id")
            relation_ids.add(relation["relation_id"])
            relations_by_id[relation["relation_id"]] = relation
            if relation["source_issue_id"] not in issue_ids or relation["target_issue_id"] not in issue_ids: fail("IssueRelations foreign key violation")
            if relation["source_issue_id"] == relation["target_issue_id"]: fail("self issue relation is forbidden")
            if schema_version >= LEDGER_SCHEMA_VERSION and relation["relation_type"] not in RELATION_TYPES: fail(f"unknown issue relation type: {relation['relation_type']}")
        relation_events: dict[str, str] = {}; legacy_related_events: list[dict[str, Any]] = []
        for event in events:
            payload = event_payloads[event["event_id"]]
            strict_event = strict_events[event["event_id"]]
            expected_type = CONSOLIDATION_RELATIONS.get(event["to_status"]) if event["from_status"] != event["to_status"] else None
            legacy = False
            if expected_type:
                if event["event_type"] != "RESOLVE_WITHOUT_FIX": fail("consolidation transition must use RESOLVE_WITHOUT_FIX")
                relation_id = payload.get("relation_id")
                if not relation_id:
                    if strict_event: fail("consolidation event must reference a relation_id")
                    continue
            elif event["event_type"] == "RELATED":
                relation_id = payload.get("relation_id")
                if not relation_id:
                    if strict_event: fail("RELATED event must reference a relation_id")
                    legacy = True
                    candidates = [
                        relation for relation in relations
                        if relation["source_issue_id"] == event["issue_id"]
                        and relation["created_at"] == event["occurred_at"]
                        and relation["created_by"] == event["actor"]
                    ]
                    if len(candidates) == 1:
                        relation_id = candidates[0]["relation_id"]
                    else:
                        legacy_related_events.append(event)
                        continue
            else:
                if payload.get("relation_id"): fail("non-relation event must not reference relation_id")
                continue
            relation = relations_by_id.get(str(relation_id))
            if not relation: fail("relation event references a missing relation")
            if relation["relation_id"] in relation_events: fail("relation must be referenced by exactly one event")
            if relation["source_issue_id"] != event["issue_id"]: fail("relation source does not match event issue")
            if strict_event and relation["relation_type"] not in RELATION_TYPES: fail(f"unknown issue relation type: {relation['relation_type']}")
            if expected_type and relation["relation_type"] != expected_type: fail("consolidation relation type does not match terminal status")
            if not expected_type and relation["relation_type"] in CONSOLIDATION_RELATION_TYPES and strict_event: fail("generic RELATED event cannot create a consolidation relation")
            if not legacy and payload.get("target_issue_id") != relation["target_issue_id"]: fail("relation target does not match event payload")
            if not legacy and payload.get("relation_type") != relation["relation_type"]: fail("relation type does not match event payload")
            if not legacy and relation["created_at"] != event["occurred_at"]: fail("relation created_at does not match event")
            if not legacy and relation["created_by"] != event["actor"]: fail("relation created_by does not match event")
            relation_events[relation["relation_id"]] = event["event_id"]
        unreferenced_relations = relation_ids - set(relation_events)
        if unreferenced_relations and schema_version >= LEDGER_SCHEMA_VERSION: fail("every relation must be referenced by exactly one event")
        for relation_id in unreferenced_relations:
            relation = relations_by_id[relation_id]
            if not any(
                event["issue_id"] == relation["source_issue_id"] and event["actor"] == relation["created_by"]
                for event in legacy_related_events
            ):
                fail("legacy relation has no compatible RELATED event")
        return {"valid": True, "counts": {name: len(self.rows(name)) for name in SHEETS}}

    @staticmethod
    def _relative(value: Any) -> None:
        if not value: return
        path = PurePosixPath(str(value))
        if path.is_absolute() or ".." in path.parts or "\\" in str(value): fail("relative_path must be a safe repository-relative POSIX path")


def output(value: Any) -> None:
    print(canonical(value))


def field(payload: dict[str, Any], args: argparse.Namespace, name: str, required_field: bool = False) -> Any:
    value = getattr(args, name, None)
    if value is None: value = payload.get(name)
    return required(value, name) if required_field else value


def repository_paths(values: Optional[list[str]]) -> dict[str, Path]:
    result: dict[str, Path] = {}
    for value in values or []:
        if "=" not in value: fail("--repo-path must be repo_id=/absolute/path")
        repo_id, text_path = value.split("=", 1); path = Path(text_path)
        if not repo_id or not path.is_absolute() or not path.is_dir(): fail("--repo-path must use a valid absolute directory")
        if repo_id in result: fail(f"duplicate repo mapping: {repo_id}")
        result[repo_id] = path
    return result


def git(repo: Path, *args: str) -> str:
    try:
        return subprocess.run(["git", "-C", str(repo), *args], text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=True).stdout.strip()
    except (OSError, subprocess.CalledProcessError) as exc:
        detail = getattr(exc, "stderr", "").strip()
        fail(f"git evidence validation failed: {detail or exc}")


def validate_evidence_git(ledger: Ledger, repo_id: str, commit: str, relative_path: Optional[str], mappings: dict[str, Path]) -> str:
    repo = ledger.find("Repositories", "repo_id", repo_id)
    if not repo["active"]: fail("repository is inactive")
    if repo_id not in mappings: fail(f"missing --repo-path mapping for repository: {repo_id}")
    path = mappings[repo_id]
    origin = git(path, "config", "--get", "remote.origin.url")
    if origin != repo["canonical_url"]: fail("runtime repository origin does not match canonical_url")
    canonical_commit = git(path, "rev-parse", "--verify", f"{commit}^{{commit}}")
    if not SHA.fullmatch(canonical_commit): fail("git did not return a canonical full commit hash")
    git(path, "cat-file", "-e", f"{canonical_commit}^{{commit}}")
    Ledger._relative(relative_path)
    if relative_path:
        git(path, "cat-file", "-e", f"{canonical_commit}:{relative_path}")
    return canonical_commit


def idem(ledger: Ledger, key: str) -> Optional[dict[str, Any]]:
    required(key, "idempotency_key")
    for event in ledger.rows("Events"):
        if json_object(event["payload_json"], "stored event payload").get("idempotency_key") == key:
            return event
    return None


def idempotency_fingerprint(command: str, args: argparse.Namespace, payload: dict[str, Any]) -> str:
    request: dict[str, Any] = {
        "command": command,
        "payload": {
            key: value
            for key, value in payload.items()
            if key not in {"idempotency_key", "request_fingerprint"}
        },
    }
    for name in IDEMPOTENCY_FIELDS:
        value = getattr(args, name, None)
        if value is not None:
            request[name] = value
    return hashlib.sha256(canonical(request).encode()).hexdigest()


def assert_idempotent_request(event: dict[str, Any], fingerprint: str, event_type: str, issue_id: Optional[str] = None) -> None:
    stored = json_object(event["payload_json"], "stored event payload")
    if not stored.get("request_fingerprint"):
        fail("legacy idempotency_key cannot be replayed; inspect current state and use a new key")
    conflict = event["event_type"] != event_type or (issue_id is not None and event["issue_id"] != issue_id)
    conflict = conflict or stored["request_fingerprint"] != fingerprint
    if conflict:
        fail("idempotency_key conflict: key is bound to a different request")


def append_event(ledger: Ledger, issue_id: str, event_type: str, from_status: str, to_status: str, actor: str, payload: dict[str, Any], corrects: str = "") -> dict[str, Any]:
    events = [e for e in ledger.rows("Events") if e["issue_id"] == issue_id]
    previous = events[-1]["event_hash"] if events else ""
    event = {"event_id": str(uuid.uuid4()), "issue_id": issue_id, "event_type": event_type, "from_status": from_status, "to_status": to_status, "occurred_at": now(), "actor": required(actor, "actor"), "payload_json": canonical(payload), "previous_hash": previous, "event_hash": "", "corrects_event_id": corrects or ""}
    event["event_hash"] = event_digest(event)
    ledger.append("Events", event)
    return event


def issue_mutation(ledger: Ledger, args: argparse.Namespace, action: str, target: Optional[str], updates: dict[str, Any], payload: dict[str, Any], require_authorized: bool = False, corrects: str = "") -> dict[str, Any]:
    old = ledger.find("Issues", "issue_id", required(args.issue_id, "issue_id"))
    if int(required(args.expected_version, "expected_version")) != old["version"]: fail("expected_version conflict")
    if require_authorized and (old["status"] != "READY_TO_FIX" or not old["authorized_by"] or not old["authorization_scope"] or not old["authorized_at"]): fail("fix authorization is required before start-fix")
    if target and target not in TRANSITIONS.get(old["status"], set()): fail(f"invalid state transition: {old['status']} -> {target}")
    new_status = target or old["status"]
    event = append_event(ledger, old["issue_id"], action, old["status"], new_status, field(payload, args, "actor", True), payload, corrects)
    if target: updates["status"] = target
    updates.update({"version": old["version"] + 1, "latest_event_id": event["event_id"], "updated_at": now()})
    issue = ledger.replace("Issues", "issue_id", old["issue_id"], updates)
    return {"issue": issue, "event": event}


def relation_record(ledger: Ledger, source: str, target: str, relation_type: str, actor: str, relation_id: Optional[str] = None) -> dict[str, Any]:
    ledger.find("Issues", "issue_id", source); ledger.find("Issues", "issue_id", target)
    if source == target: fail("cannot relate an issue to itself")
    if relation_type not in RELATION_TYPES: fail(f"relation_type must be one of: {', '.join(sorted(RELATION_TYPES))}")
    return {
        "relation_id": relation_id or str(uuid.uuid4()),
        "source_issue_id": source,
        "target_issue_id": target,
        "relation_type": relation_type,
        "created_at": now(),
        "created_by": required(actor, "actor"),
    }


def relation_for_event(ledger: Ledger, event: dict[str, Any]) -> dict[str, Any]:
    payload = json_object(event["payload_json"], "stored event payload")
    if payload.get("relation_id"):
        return ledger.find("IssueRelations", "relation_id", payload["relation_id"])
    matches = [
        relation for relation in ledger.rows("IssueRelations")
        if relation["source_issue_id"] == event["issue_id"]
        and relation["created_at"] == event["occurred_at"]
        and relation["created_by"] == event["actor"]
    ]
    if len(matches) != 1:
        fail("relation event cannot be resolved to exactly one relation")
    return matches[0]


def idempotent_issue_result(ledger: Ledger, event: dict[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {
        "idempotent": True,
        "issue": ledger.find("Issues", "issue_id", event["issue_id"]),
        "event": event,
    }
    if event["event_type"] == "RELATED" or (
        event["from_status"] != event["to_status"]
        and event["to_status"] in CONSOLIDATION_RELATIONS
    ):
        result["relation"] = relation_for_event(ledger, event)
    return result


def command_init(args: argparse.Namespace) -> dict[str, Any]:
    path = Path(required(args.workbook, "workbook"))
    if path.exists(): fail("refusing to overwrite an existing workbook")
    if path.suffix.lower() != ".xlsx": fail("workbook must have .xlsx suffix")
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = json_object(args.payload_json)
    project_id = field(payload, args, "project_id", True); name = field(payload, args, "name", True)
    wb = Workbook()
    active_sheet = wb.active
    if active_sheet is None:
        fail("openpyxl did not create an active worksheet")
    wb.remove(active_sheet)
    for sheet, headers in SHEETS.items():
        ws = wb.create_sheet(sheet); ws.append(headers)
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 24
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.orientation = "landscape" if len(headers) > 6 else "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_title_rows = "1:1"
        for column, header in enumerate(headers, start=1):
            header_cell = ws.cell(1, column)
            header_cell.fill = HEADER_FILL
            header_cell.font = HEADER_FONT
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            width = TEXT_COLUMNS.get(
                header,
                FIELD_WIDTHS.get(header, max(12, min(24, len(header) + 4))),
            )
            ws.column_dimensions[header_cell.column_letter].width = width
        table = Table(displayName=sheet, ref=f"A1:{chr(64 + len(headers))}1")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    ledger = object.__new__(Ledger); ledger.path = path; ledger.writable = True; ledger.wb = wb
    ledger.append("Project", {"project_id": project_id, "name": name, "created_at": now(), "version": LEDGER_SCHEMA_VERSION})
    for state in STATES:
        ledger.append("Dictionaries", {"dictionary_type": "issue_status", "key": state, "value": state, "active": True})
    repositories = payload.get("repositories", [])
    if not isinstance(repositories, list): fail("repositories must be a JSON list")
    for item in repositories:
        if not isinstance(item, dict) or set(item) != set(SHEETS["Repositories"]): fail("repository definitions must contain exactly the Repositories fields")
        if "local_path" in item: fail("local_path must never be persisted")
        ledger.append("Repositories", item)
    ledger.save()
    return {"created": str(path), "project_id": project_id}


def command_report(ledger: Ledger, args: argparse.Namespace) -> dict[str, Any]:
    payload = json_object(args.payload_json); key = required(args.idempotency_key, "idempotency_key")
    title = field(payload, args, "title", True); description = field(payload, args, "description", True); actor = field(payload, args, "actor", True)
    severity = field(payload, args, "severity", True)
    fingerprint = idempotency_fingerprint("issue:report", args, payload)
    existing = idem(ledger, key)
    if existing:
        assert_idempotent_request(existing, fingerprint, "REPORTED")
        return {"idempotent": True, "issue": ledger.find("Issues", "issue_id", existing["issue_id"]), "event": existing}
    project = ledger.rows("Project")[0]
    pattern = re.compile(re.escape(project["project_id"]) + r"-BUG-(\d{4})$")
    sequence = max((int(match.group(1)) for row in ledger.rows("Issues") if (match := pattern.fullmatch(str(row["issue_id"])))), default=0) + 1
    issue_id = f"{project['project_id']}-BUG-{sequence:04d}"
    issue = {"issue_id": issue_id, "project_id": project["project_id"], "title": title, "description": description, "severity": severity, "status": "REPORTED", "version": 1, "latest_event_id": "", "reported_at": now(), "reported_by": actor, "root_cause": "", "authorization_scope": "", "authorized_by": "", "authorized_at": "", "resolution": "", "closed_at": "", "updated_at": now()}
    ledger.append("Issues", issue); payload.update({"idempotency_key": key, "request_fingerprint": fingerprint, "title": title, "description": description, "severity": severity})
    event = append_event(ledger, issue_id, "REPORTED", "REPORTED", "REPORTED", actor, payload)
    issue = ledger.replace("Issues", "issue_id", issue_id, {"latest_event_id": event["event_id"]})
    return {"issue": issue, "event": event}


def command_transition(ledger: Ledger, args: argparse.Namespace, name: str) -> dict[str, Any]:
    payload = json_object(args.payload_json); key = required(args.idempotency_key, "idempotency_key")
    issue_id = required(args.issue_id, "issue_id"); required(args.expected_version, "expected_version")
    actor = field(payload, args, "actor", True)
    event_type = name.upper().replace("-", "_")
    fingerprint = idempotency_fingerprint(f"issue:{name}", args, payload)
    existing = idem(ledger, key)
    if existing:
        requested_resolution = field(payload, args, "resolution", True) if name == "resolve-without-fix" else None
        assert_idempotent_request(existing, fingerprint, event_type, issue_id)
        if requested_resolution is not None and existing["to_status"] != requested_resolution:
            fail("idempotency_key conflict: key is bound to a different request")
        result = idempotent_issue_result(ledger, existing)
        if requested_resolution in CONSOLIDATION_RELATIONS:
            existing_relation: dict[str, Any] = result["relation"]
            requested_target = field(payload, args, "target_issue_id", True)
            if existing_relation["target_issue_id"] != requested_target or existing_relation["relation_type"] != CONSOLIDATION_RELATIONS[requested_resolution]:
                fail("idempotency_key conflict: key is bound to a different request")
        return result
    payload.update({"idempotency_key": key, "request_fingerprint": fingerprint})
    target: Optional[str] = None; updates: dict[str, Any] = {}; relation: Optional[dict[str, Any]] = None
    if name == "record-history-search": target = "TRIAGED"
    elif name == "confirm": target = "CONFIRMED"
    elif name == "start-investigation": target = "INVESTIGATING"
    elif name == "identify-root-cause":
        root_cause = field(payload, args, "root_cause", True); updates["root_cause"] = root_cause; payload["root_cause"] = root_cause
    elif name == "authorize-fix":
        authorized_by = field(payload, args, "authorized_by", True); authorization_scope = field(payload, args, "authorization_scope", True)
        target = "READY_TO_FIX"; updates.update({"authorized_by": authorized_by, "authorized_at": now(), "authorization_scope": authorization_scope}); payload.update({"authorized_by": authorized_by, "authorization_scope": authorization_scope})
    elif name == "revoke-fix-authorization": target = "INVESTIGATING"; updates.update({"authorized_by": "", "authorized_at": "", "authorization_scope": ""})
    elif name == "start-fix": target = "FIXING"
    elif name == "record-progress":
        current = ledger.find("Issues", "issue_id", issue_id)
        if current["status"] in TERMINAL: fail(f"record-progress is forbidden for terminal issue status: {current['status']}")
        progress = field(payload, args, "progress", True); payload["progress"] = progress
    elif name == "complete-fix":
        if not any(e["issue_id"] == args.issue_id and e["evidence_type"] == "FIX_COMMIT" and e["active"] in (True, 1) for e in ledger.rows("Evidence")): fail("complete-fix requires active FIX_COMMIT evidence")
        target = "READY_TO_VERIFY"
    elif name == "record-verification":
        if not any(e["issue_id"] == args.issue_id and e["evidence_type"] == "VERIFICATION" and e["active"] in (True, 1) for e in ledger.rows("Evidence")): fail("record-verification requires active VERIFICATION evidence")
        outcome = field(payload, args, "verification_result", True)
        if outcome not in {"PASS", "FAIL"}: fail("verification_result must be PASS or FAIL")
        target = "CLOSED" if outcome == "PASS" else "REOPENED"
        verification_resolution = field(payload, args, "resolution", True)
        updates.update({"resolution": verification_resolution, "closed_at": now() if outcome == "PASS" else ""}); payload.update({"verification_result": outcome, "resolution": verification_resolution})
    elif name == "reopen": target = "REOPENED"; updates["closed_at"] = ""
    elif name == "defer":
        defer_resolution = field(payload, args, "resolution", True); target = "DEFERRED"; updates["resolution"] = defer_resolution; payload["resolution"] = defer_resolution
    elif name == "resolve-without-fix":
        target = field(payload, args, "resolution", True)
        if target not in {"DUPLICATE", "MERGED", "INVALID", "CANNOT_REPRODUCE", "WONT_FIX"}: fail("resolution must be a terminal non-fix state")
        payload["resolution"] = target
        if target in CONSOLIDATION_RELATIONS:
            source = issue_id
            relation = relation_record(
                ledger,
                source,
                field(payload, args, "target_issue_id", True),
                CONSOLIDATION_RELATIONS[target],
                actor,
            )
            payload.update({
                "relation_id": relation["relation_id"],
                "target_issue_id": relation["target_issue_id"],
                "relation_type": relation["relation_type"],
            })
        updates["resolution"] = target; updates["closed_at"] = now()
    else: fail(f"unsupported issue command: {name}")
    result = issue_mutation(ledger, args, event_type, target, updates, payload, name == "start-fix")
    if relation:
        relation["created_at"] = result["event"]["occurred_at"]
        relation["created_by"] = result["event"]["actor"]
        ledger.append("IssueRelations", relation)
        result["relation"] = relation
    return result


def command_relate(ledger: Ledger, args: argparse.Namespace) -> dict[str, Any]:
    payload = json_object(args.payload_json); key = required(args.idempotency_key, "idempotency_key")
    source = required(args.issue_id, "issue_id"); required(args.expected_version, "expected_version")
    target = field(payload, args, "target_issue_id", True); relation_type = field(payload, args, "relation_type", True)
    actor = field(payload, args, "actor", True)
    if relation_type in CONSOLIDATION_RELATION_TYPES: fail("DUPLICATE_OF and MERGED_INTO must be created by resolve-without-fix")
    fingerprint = idempotency_fingerprint("issue:relate", args, payload)
    existing = idem(ledger, key)
    if existing:
        assert_idempotent_request(existing, fingerprint, "RELATED", source)
        result = idempotent_issue_result(ledger, existing)
        if result["relation"]["target_issue_id"] != target or result["relation"]["relation_type"] != relation_type:
            fail("idempotency_key conflict: key is bound to a different request")
        return result
    relation = relation_record(ledger, source, target, relation_type, actor)
    payload.update({
        "idempotency_key": key,
        "request_fingerprint": fingerprint,
        "relation_id": relation["relation_id"],
        "target_issue_id": relation["target_issue_id"],
        "relation_type": relation["relation_type"],
    })
    result = issue_mutation(ledger, args, "RELATED", None, {}, payload)
    relation["created_at"] = result["event"]["occurred_at"]
    relation["created_by"] = result["event"]["actor"]
    ledger.append("IssueRelations", relation)
    result["relation"] = relation
    return result


def command_evidence(ledger: Ledger, args: argparse.Namespace, mappings: dict[str, Path]) -> dict[str, Any]:
    payload = json_object(args.payload_json); key = required(args.idempotency_key, "idempotency_key")
    event_type = "EVIDENCE_REGISTERED" if args.evidence_command == "register" else "EVIDENCE_RETRACTED"
    fingerprint = idempotency_fingerprint(f"evidence:{args.evidence_command}", args, payload)
    existing = idem(ledger, key)
    if existing:
        assert_idempotent_request(existing, fingerprint, event_type, field(payload, args, "issue_id"))
        return {"idempotent": True, "event": existing}
    payload["request_fingerprint"] = fingerprint
    if args.evidence_command == "register":
        args.issue_id = field(payload, args, "issue_id", True)
        repo_id = field(payload, args, "repo_id", True); relative_path = field(payload, args, "relative_path")
        commit = validate_evidence_git(ledger, repo_id, field(payload, args, "commit", True), relative_path, mappings)
        evidence_type = field(payload, args, "evidence_type", True); description = field(payload, args, "description", True)
        payload.update({"idempotency_key": key, "evidence_type": evidence_type, "repo_id": repo_id, "commit_hash": commit, "relative_path": relative_path or "", "description": description})
        result = issue_mutation(ledger, args, "EVIDENCE_REGISTERED", None, {}, payload)
        evidence = {"evidence_id": str(uuid.uuid4()), "issue_id": args.issue_id, "evidence_type": evidence_type, "repo_id": repo_id, "commit_hash": commit, "relative_path": relative_path or "", "locator": field(payload, args, "locator") or "", "description": description, "registered_at": now(), "registered_by": field(payload, args, "actor", True), "active": True}
        ledger.append("Evidence", evidence); ledger.append("EventEvidence", {"event_id": result["event"]["event_id"], "evidence_id": evidence["evidence_id"]})
        result["evidence"] = evidence; return result
    evidence = ledger.find("Evidence", "evidence_id", required(args.evidence_id, "evidence_id")); args.issue_id = evidence["issue_id"]
    payload["idempotency_key"] = key; payload["evidence_id"] = evidence["evidence_id"]
    result = issue_mutation(ledger, args, "EVIDENCE_RETRACTED", None, {}, payload)
    evidence = ledger.replace("Evidence", "evidence_id", evidence["evidence_id"], {"active": False})
    ledger.append("EventEvidence", {"event_id": result["event"]["event_id"], "evidence_id": evidence["evidence_id"]})
    result["evidence"] = evidence; return result


def command_correct(ledger: Ledger, args: argparse.Namespace) -> dict[str, Any]:
    payload = json_object(args.payload_json); key = required(args.idempotency_key, "idempotency_key")
    fingerprint = idempotency_fingerprint("event:correct", args, payload)
    existing = idem(ledger, key)
    if existing:
        assert_idempotent_request(existing, fingerprint, "EVENT_CORRECTED")
        return {"idempotent": True, "event": existing}
    original = ledger.find("Events", "event_id", required(args.event_id, "event_id")); args.issue_id = original["issue_id"]
    payload["idempotency_key"] = key; payload["request_fingerprint"] = fingerprint; payload["corrected_event_id"] = original["event_id"]
    result = issue_mutation(ledger, args, "EVENT_CORRECTED", None, {}, payload, corrects=original["event_id"])
    # Correcting immutable history is represented only by the appended event.
    return result


def parser() -> argparse.ArgumentParser:
    root = argparse.ArgumentParser(description=__doc__)
    root.add_argument("--workbook", help="explicit .xlsx ledger path")
    root.add_argument("--repo-path", action="append", help="runtime repo_id=/absolute/path mapping; never persisted")
    groups = root.add_subparsers(dest="group", required=True)
    ledger = groups.add_parser("ledger"); lp = ledger.add_subparsers(dest="command", required=True)
    init = lp.add_parser("init"); init.add_argument("--project-id"); init.add_argument("--name"); init.add_argument("--payload-json")
    lp.add_parser("validate"); lp.add_parser("project-info")
    issue = groups.add_parser("issue"); ip = issue.add_subparsers(dest="issue_command", required=True)
    report = ip.add_parser("report"); report.add_argument("--title"); report.add_argument("--description"); report.add_argument("--severity"); report.add_argument("--actor"); report.add_argument("--payload-json"); report.add_argument("--idempotency-key")
    get = ip.add_parser("get"); get.add_argument("--issue-id")
    ip.add_parser("list")
    search = ip.add_parser("search"); search.add_argument("--query")
    history = ip.add_parser("history"); history.add_argument("--issue-id")
    relations = ip.add_parser("relations"); relations.add_argument("--issue-id")
    relation_get = ip.add_parser("relation-get"); relation_get.add_argument("--relation-id")
    changing = {
        "record-history-search": (),
        "confirm": (),
        "start-investigation": (),
        "identify-root-cause": ("root_cause",),
        "authorize-fix": ("authorized_by", "authorization_scope"),
        "revoke-fix-authorization": (),
        "start-fix": (),
        "record-progress": ("progress",),
        "complete-fix": (),
        "record-verification": ("resolution", "verification_result"),
        "reopen": (),
        "defer": ("resolution",),
        "resolve-without-fix": ("resolution",),
    }
    for name, command_fields in changing.items():
        p = ip.add_parser(name); p.add_argument("--issue-id"); p.add_argument("--expected-version", type=int); p.add_argument("--actor"); p.add_argument("--idempotency-key"); p.add_argument("--payload-json")
        for field_name in command_fields:
            p.add_argument("--" + field_name.replace("_", "-"))
        if name == "resolve-without-fix": p.add_argument("--target-issue-id")
    relate = ip.add_parser("relate"); relate.add_argument("--issue-id"); relate.add_argument("--target-issue-id"); relate.add_argument("--relation-type"); relate.add_argument("--expected-version", type=int); relate.add_argument("--actor"); relate.add_argument("--idempotency-key"); relate.add_argument("--payload-json")
    evidence = groups.add_parser("evidence"); ep = evidence.add_subparsers(dest="evidence_command", required=True)
    register = ep.add_parser("register")
    for name in ("issue_id", "evidence_type", "repo_id", "commit", "relative_path", "locator", "description", "actor", "idempotency_key"):
        register.add_argument("--" + name.replace("_", "-"))
    register.add_argument("--expected-version", type=int); register.add_argument("--payload-json")
    retract = ep.add_parser("retract"); retract.add_argument("--evidence-id"); retract.add_argument("--expected-version", type=int); retract.add_argument("--actor"); retract.add_argument("--idempotency-key"); retract.add_argument("--payload-json")
    event = groups.add_parser("event"); evp = event.add_subparsers(dest="event_command", required=True)
    correct = evp.add_parser("correct"); correct.add_argument("--event-id"); correct.add_argument("--expected-version", type=int); correct.add_argument("--actor"); correct.add_argument("--idempotency-key"); correct.add_argument("--payload-json")
    return root


def normalize_global_args(argv: list[str]) -> list[str]:
    """Permit explicit global runtime arguments before or after a subcommand."""
    global_args: list[str] = []; rest: list[str] = []; i = 0
    while i < len(argv):
        if argv[i] in {"--workbook", "--repo-path"}:
            if i + 1 == len(argv): fail(f"{argv[i]} requires a value")
            global_args.extend(argv[i:i + 2]); i += 2
        else: rest.append(argv[i]); i += 1
    return global_args + rest


def main(argv: Optional[list[str]] = None) -> int:
    try:
        args = parser().parse_args(normalize_global_args(list(sys.argv[1:] if argv is None else argv)))
        if args.group == "ledger" and args.command == "init":
            result = command_init(args); output({"ok": True, "result": result}); return 0
        path = Path(required(args.workbook, "workbook"))
        writable = not (args.group == "ledger" and args.command in {"validate", "project-info"}) and not (args.group == "issue" and args.issue_command in {"get", "list", "search", "history", "relations", "relation-get"})
        with workbook_lock(path, writable):
            ledger = Ledger(path, writable); ledger.validate()
            if args.group == "ledger": result = ledger.validate() if args.command == "validate" else {"project": ledger.rows("Project")[0], "repositories": ledger.rows("Repositories")}
            elif args.group == "issue":
                if args.issue_command == "get": result = ledger.find("Issues", "issue_id", required(args.issue_id, "issue_id"))
                elif args.issue_command == "list": result = ledger.rows("Issues")
                elif args.issue_command == "search":
                    query = required(args.query, "query").casefold(); result = [r for r in ledger.rows("Issues") if query in canonical(r).casefold()]
                elif args.issue_command == "history":
                    issue_id = required(args.issue_id, "issue_id"); ledger.find("Issues", "issue_id", issue_id); result = [e for e in ledger.rows("Events") if e["issue_id"] == issue_id]
                elif args.issue_command == "relations":
                    issue_id = required(args.issue_id, "issue_id"); ledger.find("Issues", "issue_id", issue_id); result = [r for r in ledger.rows("IssueRelations") if issue_id in {r["source_issue_id"], r["target_issue_id"]}]
                elif args.issue_command == "relation-get": result = ledger.find("IssueRelations", "relation_id", required(args.relation_id, "relation_id"))
                elif args.issue_command == "relate": result = command_relate(ledger, args)
                elif args.issue_command == "report": result = command_report(ledger, args)
                else: result = command_transition(ledger, args, args.issue_command)
            elif args.group == "evidence": result = command_evidence(ledger, args, repository_paths(args.repo_path))
            else: result = command_correct(ledger, args)
            if writable and not (isinstance(result, dict) and result.get("idempotent")):
                ledger.validate(); ledger.save()
            output({"ok": True, "result": result})
        return 0
    except LedgerError as exc:
        output({"ok": False, "error": str(exc)})
        return 2
    except Exception as exc:
        output({"ok": False, "error": f"unexpected failure: {exc.__class__.__name__}: {exc}"})
        return 3


if __name__ == "__main__":
    sys.exit(main())
